from main_form import *
from ex_edit_form import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import xlrd
import openpyxl


class MainForm(QMainWindow, Ui_main_form):
    __style_unclicked='''QPushButton
{
	border-color: rgb(255, 255, 255);
	color: rgb(255, 255, 255);
}
QPushButton:hover
{
	background-color: rgb(14, 92, 47);
}
QPushButton:pressed
{
	color: rgb(33, 115, 70);
	background-color: rgb(243, 243, 243);
}'''

    __style_clicked='''QPushButton
{
	color: rgb(33, 115, 70);
	background-color: rgb(243, 243, 243);
}'''

    def __init__(self, parent=None):
        super(MainForm, self).__init__(parent)
        self.setupUi(self)
        self.is_click_search = False
        self.statusbar.showMessage("就绪")
        self.btn_search.clicked.connect(self.btn_search_clicked)
        self.btn_add.clicked.connect(self.btn_add_clicked)
        self.btn_save.clicked.connect(self.btn_save_clicked)
        self.btn_begin_search.clicked.connect(self.btn_begin_search_clicked)
        self.btn_del.clicked.connect(self.btn_del_clicked)
        self.le_keyword.returnPressed.connect(self.btn_begin_search_clicked)
        self.table_data = None
        self.search_list = ["", 0, object]
        self.init_data_book()
        self.init_table()

    def btn_del_clicked(self):
        items = self.tw_content.selectedItems()
        if len(items) != 0:
            self.tw_content.removeRow(items[0].row())

    def btn_search_clicked(self):
        if self.is_click_search:
            self.btn_search.setStyleSheet(self.__style_unclicked)
            self.wg_search.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Ignored)
        else:
            self.btn_search.setStyleSheet(self.__style_clicked)
            self.wg_search.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)
            self.le_keyword.setFocus()
        self.is_click_search = not self.is_click_search

    def btn_add_clicked(self):
        editForm = EditForm(self)
        editForm.setWindowModality(Qt.ApplicationModal)
        editForm.is_edit = False
        editForm.target_row = self.tw_content.rowCount()
        editForm.show()

    def btn_save_clicked(self):
        try:
            self.save()
            info = "已保存"
        except PermissionError:
            info = "保存失败，文件正在占用或所在用户权限不足"
        QMessageBox.information(self, "提示", info)

    def btn_begin_search_clicked(self):
        self.remove_selected()
        if len(self.le_keyword.text()) != 0:
            self.highlight_search_result()
        else:
            self.remove_highlight()

    def tw_content_cellDoubleClicked(self, row_index, column_index):
        edit_form = EditForm(self)
        edit_form.list = [self.tw_content.item(row_index, i).text() for i in range(6)]
        edit_form.target_row = row_index
        edit_form.is_edit = True
        edit_form.init_list()
        edit_form.show()

    def highlight_search_result(self):
        try:
            if self.search_list[0] != self.le_keyword.text():
                item = self.find_new_key()
            else:
                item = self.find_old_key()
            row = item.row()
            self.tw_content.verticalScrollBar().setSliderPosition(row)
            self.tw_content.selectRow(row)
        except ValueError:
            QMessageBox.information(self, "提示", "没有搜索到结果")

    def find_new_key(self):
        self.remove_highlight()
        items = self.tw_content.findItems(self.le_keyword.text(), Qt.MatchContains)
        if len(items) == 0:
            raise ValueError
        self.search_list[0] = self.le_keyword.text()
        self.search_list[1] = 0
        self.search_list[2] = items
        for i in items:
            i.setBackground(QBrush(QColor(255, 238, 128)))
        return items[0]

    def find_old_key(self):
        if self.search_list[1] == len(self.search_list[2]) - 1:
            self.search_list[1] = 0
        else:
            self.search_list[1] += 1
        return self.search_list[2][self.search_list[1]]

    def remove_highlight(self):
        for i in range(self.tw_content.rowCount()):
            for j in range(self.tw_content.columnCount()):
                self.tw_content.item(i, j).setBackground(QBrush(QColor(255, 255, 255)))

    def remove_selected(self):
        for item in self.tw_content.selectedItems():
            item.setSelected(False)

    def add_one_row(self, row_list):
        row_index = self.tw_content.rowCount()
        self.tw_content.setRowCount(row_index + 1)
        self.edit_row(row_index, row_list)

    def edit_row(self, pos, row_list):
        for i in range(len(row_list)):
            new_item = QTableWidgetItem(row_list[i])
            self.tw_content.setItem(pos, i, new_item)

    def init_data_book(self):
        path = 'info_class.xlsx'
        try:
            book = xlrd.open_workbook(path)
        except FileNotFoundError:
            QMessageBox.information(self, "错误", "未找到文件，请确认数据文件与程序在同一个目录中")
            self.close()
        self.lb_title.setText(path+" - Excel")
        self.table_data = book.sheets()[0]

    def init_table(self):
        self.tw_content.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tw_content.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tw_content.setRowCount(self.table_data.nrows)
        self.tw_content.setColumnCount(self.table_data.ncols)
        self.tw_content.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tw_content.cellDoubleClicked.connect(self.tw_content_cellDoubleClicked)
        self.init_table_data()

    def init_table_data(self):
        for i in range(self.table_data.nrows):
            for j in range(self.table_data.ncols):
                if type(self.table_data.cell(i, j).value) == float:
                    val = str(int(self.table_data.cell(i, j).value))
                else:
                    val = str(self.table_data.cell(i, j).value)
                new_item = QTableWidgetItem(val)
                self.tw_content.setItem(i, j, new_item)
        self.tw_content.setHorizontalHeaderLabels(['学号', '姓名', '性别', '年龄', '联系电话', '家庭住址'])

    def save(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for i in range(self.tw_content.rowCount()):
            for j in range(self.tw_content.columnCount()):
                cell = sheet.cell(i+1, j+1)
                cell.value = self.tw_content.item(i, j).text()
        workbook.save('info_class.xlsx')
